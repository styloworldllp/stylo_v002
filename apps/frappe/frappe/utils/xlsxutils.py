# Copyright (c) 2015, Frappe Technologies Pvt. Ltd. and Contributors
# License: MIT. See LICENSE
import datetime
import functools
import re
from dataclasses import dataclass
from dataclasses import field as dataclass_field
from io import BytesIO
from typing import Any, ClassVar, Literal

import xlrd
import xlsxwriter
from openpyxl import load_workbook
from xlsxwriter.format import Format

import frappe
from frappe import _
from frappe.core.utils import html2text
from frappe.utils import cint
from frappe.utils.html_utils import unescape_html

ILLEGAL_CHARACTERS_RE = re.compile(
	r"[\000-\010]|[\013-\014]|[\016-\037]|\uFEFF|\uFFFE|\uFFFF|[\uD800-\uDFFF]"
)

# as required by XLSXWriter
INVALID_SHEET_NAME_RE = re.compile(r"[\[\]:*?/\\]")
MAX_SHEET_NAME_LENGTH = 31


### XLSX Formatter ###
@dataclass(slots=True)
class XLSXMetadata:
	"""
	Metadata container for XLSX report styling.

	- All indexes must be 0-based respecting xlsxwriter's indexing.

	Attributes:
		column_map: Maps column index to column dict (fieldname, fieldtype, etc.).
		row_map: Maps row index to row data (dict or list).
		applied_filters_map: Maps row index to list of applied filter label-value pairs.
		has_total_row: Whether the last row is a total row.
		has_indentation: Whether indentation styling should be applied.

		# optional metadata for custom style builders
		report_name: Name of the report.
		filters: Raw filter values.
	"""

	column_map: dict[int, dict] = dataclass_field(default_factory=dict)
	row_map: dict[int, dict | list] = dataclass_field(default_factory=dict)
	applied_filters_map: dict[int, list] = dataclass_field(default_factory=dict)

	has_total_row: bool = False
	has_indentation: bool = False

	# optional
	report_name: str = ""
	filters: dict = dataclass_field(default_factory=dict)

	def get_column(self, fieldname: str) -> dict | None:
		"""
		Get column dict by fieldname, or None if not found.
		"""
		return next((col for col in self.column_map.values() if col.get("fieldname") == fieldname), None)

	def get_header_index(self) -> int:
		"""
		Get header row index based on applied filters.
		Assumes header is always 1 row after the last filter row.
		"""
		count = len(self.applied_filters_map)
		return count + 1 if count else 0

	def get_first_row_index(self) -> int:
		return min(self.row_map.keys()) if self.row_map else 0

	def get_last_row_index(self) -> int:
		return max(self.row_map.keys()) if self.row_map else 0


class XLSXStyleBuilder:
	"""
	Builder for XLSX cell styles based on report metadata.

	Builds a style dictionary with:
		- styles: List of style definitions (xlsxwriter format properties). List index is the style ID.
		- column_styles: Maps column index to list of style IDs.
		- row_styles: Maps row index to list of style IDs.
		- cell_styles: Maps (row, col) tuple to list of style IDs.

	**Usage:**

	```
	builder = XLSXStyleBuilder(metadata)
	builder.style_column(0, builder.register_style({"bold": True}))
	styles = builder.result
	```
	"""

	RIGHT_ALIGN_FIELDTYPES: ClassVar[set[str]] = {
		*frappe.model.numeric_fieldtypes,
		*frappe.model.datetime_fields,
		"Rating",
	}

	def __init__(self, metadata: XLSXMetadata, default_styling: bool = True):
		self.metadata = metadata

		# column fieldname -> index mapping
		self.field_index_map = {
			col["fieldname"]: idx for idx, col in self.metadata.column_map.items() if col.get("fieldname")
		}

		self.styles: list[dict] = []
		self.column_styles: dict[int, list[int]] = {}
		self.row_styles: dict[int, list[int]] = {}
		self.cell_styles: dict[tuple[int, int], list[int]] = {}

		self.result = {
			"styles": self.styles,
			"column_styles": self.column_styles,
			"row_styles": self.row_styles,
			"cell_styles": self.cell_styles,
		}

		# metadata indexes for quick access
		self.header_index = self.metadata.get_header_index()
		self.first_row_index = self.metadata.get_first_row_index()
		self.last_row_index = self.metadata.get_last_row_index()
		self.row_is_dict = isinstance(self.metadata.row_map.get(self.first_row_index), dict)

		self._register_common_styles()

		if default_styling:
			self.apply_default_styles()

	### STYLE REGISTRATION ###
	def _register_common_styles(self):
		self.bold_style_id = self.register_style({"bold": True})

	def register_style(self, style: dict) -> int:
		"""
		Register a style and return its ID.

		Style dict uses xlsxwriter format properties.
		"""
		if not style:
			frappe.throw(_("Cannot register an empty XLSX style"))

		style_id = len(self.styles)
		self.styles.append(style)

		return style_id

	### STYLE APPLICATION ###
	def style_column(self, col_idx: int, style_id: int):
		"""
		Apply a style to all cells in a column.

		Args:
			col_idx: 0-based column index
			style_id: ID of the style to apply (from register_style)
		"""
		if col_idx not in self.column_styles:
			self.column_styles[col_idx] = []

		self.column_styles[col_idx].append(style_id)

		return self

	def style_row(self, row_idx: int, style_id: int):
		"""
		Apply a style to all cells in a row.

		Args:
			row_idx: 0-based row index
			style_id: ID of the style to apply (from register_style)
		"""
		if row_idx not in self.row_styles:
			self.row_styles[row_idx] = []

		self.row_styles[row_idx].append(style_id)

		return self

	def style_cell(self, row_idx: int, col_idx: int, style_id: int):
		"""
		Apply a style to a specific cell.

		Args:
			row_idx: 0-based row index
			col_idx: 0-based column index
			style_id: ID of the style to apply (from register_style)
		"""
		key = (row_idx, col_idx)
		cell_styles = self.cell_styles

		if key not in cell_styles:
			cell_styles[key] = []

		cell_styles[key].append(style_id)

		return self

	### UTILITY METHODS FOR STYLING ###
	def apply_default_styles(self, currency_formatting: bool = True):
		"""
		Apply all default styles:

		- Header row styling
		- Filter rows styling
		- Total row styling (if has_total_row)
		- Indentation styling (if has_indentation)
		- Default fieldtype formatting (numbers, dates, etc.)
			- Currency formatting can be toggled with currency_formatting flag
		"""
		self.style_header()

		if self.metadata.applied_filters_map:
			self.style_filters()

		if self.metadata.has_total_row:
			self.style_total_row()

		if self.metadata.has_indentation:
			self.apply_indentations()

		self.apply_default_fieldtype_formats(currency_formatting)

		return self

	def style_header(self):
		header_index = self.header_index

		self.style_row(header_index, self.bold_style_id)

		right_align = self.register_style({"align": "right"})
		left_align = self.register_style({"align": "left"})

		for col_idx, col in self.metadata.column_map.items():
			self.style_cell(
				header_index,
				col_idx,
				right_align if col.get("fieldtype") in self.RIGHT_ALIGN_FIELDTYPES else left_align,
			)

		return self

	def style_filters(self):
		for row_idx in self.metadata.applied_filters_map.keys():
			# style only the label column (0th index)
			self.style_cell(row_idx, 0, self.bold_style_id)
		return self

	def apply_indentations(self, col_idx: int = 0, field: str = "indent", pt: int = 2):
		if not self.row_is_dict:
			return self

		@functools.cache
		def register_indent_style(indent: int) -> int:
			return self.register_style({"align": "left", "indent": indent * pt})

		# quick access for hot loop
		last_row_index = self.last_row_index
		skip_last_row = self.metadata.has_total_row
		style_cell = self.style_cell

		for row_idx, row in self.metadata.row_map.items():
			if skip_last_row and row_idx == last_row_index:
				continue

			if indent := row.get(field):
				style_cell(row_idx, col_idx, register_indent_style(indent))

		return self

	def style_total_row(self):
		return self.style_row(self.last_row_index, self.bold_style_id)

	def apply_default_fieldtype_formats(self, currency_formatting: bool = True):
		formats: dict[str, int] = {
			"Float": self.register_style({"num_format": self.get_number_format("Float")}),
			"Percent": self.register_style({"num_format": self.get_number_format("Percent")}),
			"Date": self.register_style({"num_format": self.get_date_format(), "align": "right"}),
			"Time": self.register_style({"num_format": self.get_time_format(), "align": "right"}),
			"Datetime": self.register_style({"num_format": self.get_datetime_format(), "align": "right"}),
		}

		for idx, col in self.metadata.column_map.items():
			style_id = formats.get(col.get("fieldtype"))

			if style_id is not None:
				self.style_column(idx, style_id)

		if currency_formatting:
			self.apply_currency_fieldtype_formats()

		return self

	def apply_currency_fieldtype_formats(self):
		currency_options = {
			col_idx: col.get("options")
			for col_idx, col in self.metadata.column_map.items()
			if col.get("fieldtype") == "Currency"
		}

		if not currency_options:
			return self

		default_currency = frappe.db.get_default("currency")

		# quick access for hot loop
		last_row_index = self.last_row_index
		skip_last_row = self.metadata.has_total_row
		currency_options_items = currency_options.items()
		style_cell = self.style_cell

		# helpers
		@functools.cache
		def _get_value(doctype: str, docname: str, fieldname: str) -> str | None:
			return frappe.db.get_value(doctype, docname, fieldname)

		@functools.cache
		def parse_options(options: str) -> tuple:
			parts = options.split(":")
			return parts if len(parts) == 3 else (None, None, None)

		@functools.cache
		def register_currency_style(currency: str) -> int:
			return self.register_style({"num_format": self.get_number_format("Currency", currency)})

		# dispatch dict/list row access once, not per cell
		if self.row_is_dict:

			def get_row_value(row, field):
				return row.get(field)
		else:
			_field_index_get = self.field_index_map.get

			def get_row_value(row, field):
				idx = _field_index_get(field)
				return row[idx] if idx is not None else None

		# currency formatting
		for row_idx, row in self.metadata.row_map.items():
			if skip_last_row and row_idx == last_row_index:
				continue

			for col_idx, options in currency_options_items:
				currency = None

				if options:
					if ":" not in options:
						currency = get_row_value(row, options)
					else:
						doctype, link_field, currency_field = parse_options(options)
						if doctype is not None and (link_value := get_row_value(row, link_field)):
							currency = _get_value(doctype, link_value, currency_field)

				style_cell(row_idx, col_idx, register_currency_style(currency or default_currency))

		return self

	@staticmethod
	def _get_currency_symbol_info(currency: str | None) -> tuple[str, bool]:
		if not currency or frappe.db.get_default("hide_currency_symbol") == "Yes":
			return "", False

		symbol, on_right = frappe.db.get_value("Currency", currency, ["symbol", "symbol_on_right"])

		return (symbol or currency), bool(on_right)

	@staticmethod
	def _build_currency_format(
		format_string: str,
		currency_symbol: str | None = None,
		symbol_on_right: bool = False,
	) -> str:
		if not currency_symbol:
			return format_string

		if symbol_on_right:
			return f'{format_string}" {currency_symbol}";-{format_string}" {currency_symbol}"'

		return f'"{currency_symbol} "{format_string};"{currency_symbol} "-{format_string}'

	### FORMAT GETTERS ###
	@staticmethod
	def get_number_format(
		fieldtype: Literal["Currency", "Float", "Percent"],
		currency: str | None = None,
	) -> str:
		"""
		Get Excel number format string for the given fieldtype.
		"""
		from frappe.locale import get_number_format as _get_format

		number_format = _get_format()
		thousands_sep = number_format.thousands_separator
		precision = number_format.precision

		if fieldtype == "Currency":
			precision = cint(frappe.db.get_default("currency_precision")) or precision
			format_str = XLSXStyleBuilder._build_number_format(thousands_sep, precision)
			currency_symbol, symbol_on_right = XLSXStyleBuilder._get_currency_symbol_info(currency)
			return XLSXStyleBuilder._build_currency_format(format_str, currency_symbol, symbol_on_right)

		elif fieldtype in ("Float", "Percent"):
			precision = cint(frappe.db.get_default("float_precision")) or precision
			format_str = XLSXStyleBuilder._build_number_format(thousands_sep, precision)
			return f'{format_str}"%" ' if fieldtype == "Percent" else format_str

		return "General"

	@staticmethod
	def _build_number_format(thousands_sep: str, precision: int = 0) -> str:
		# Decimal separator is hardcoded to '.' because Excel only understands '.' in format strings.
		# The system decimal separator is intentionally ignored here.
		# TODO: can be improved by passing a language/locale to xlsxwriter's Workbook for locale-aware formatting.
		integer_part = "#,##0" if thousands_sep else "#0"
		decimal_part = ("." + "0" * precision) if precision > 0 else ""

		return f"{integer_part}{decimal_part}"

	@staticmethod
	def get_date_format() -> str:
		return frappe.get_system_settings("date_format")

	@staticmethod
	def get_time_format() -> str:
		return frappe.get_system_settings("time_format")

	@staticmethod
	def get_datetime_format() -> str:
		return f"{XLSXStyleBuilder.get_date_format()} {XLSXStyleBuilder.get_time_format()}"


def get_default_xlsx_styles(
	columns: list[dict],
	data: list[list | dict],
	applied_filters: list[list] | None = None,
	*,
	has_total_row: bool = False,
	has_indentation: bool = False,
	currency_formatting: bool = True,
) -> dict:
	"""
	Generate default XLSX styles for xlsx exports.

	Args:
		columns: Column definitions with keys: fieldname, fieldtype, label, options.
		data: Row data as list of dicts or lists (excluding header and filter rows).
		applied_filters: Filter rows to display at top of sheet. Each item is [label, value].
		has_total_row: If True, applies bold styling to the last row.
		has_indentation: If True, applies indent styles based on row's 'indent' key.
		currency_formatting: If True, applies currency number formats to Currency fields.
	"""
	applied_filters = applied_filters or []
	header_index = len(applied_filters) + 1 if applied_filters else 0

	applied_filters_map = dict(enumerate(applied_filters))
	column_map = dict(enumerate(columns))
	row_map = dict(enumerate(data, start=header_index + 1))  # +1 for header row

	metadata = XLSXMetadata(
		column_map=column_map,
		row_map=row_map,
		applied_filters_map=applied_filters_map,
		has_total_row=has_total_row,
		has_indentation=has_indentation,
	)

	return XLSXStyleBuilder(metadata, default_styling=False).apply_default_styles(currency_formatting).result


### Excel Creation ###
def make_xlsx(
	data: list[list[Any]],
	sheet_name: str,
	wb: xlsxwriter.Workbook | None = None,
	column_widths: list[int] | None = None,
	styles: dict | None = None,
) -> BytesIO | None:
	"""
	Create an Excel file with the given data and formatting options.

	Args:
		data: List of rows, where each row is a list of cell values
		sheet_name: Name of the Excel sheet
		wb: Existing workbook to add sheet to. If None, creates new workbook
			- Workbook must be closed by caller if provided
			- Should be created with constant_memory=True for large datasets
		column_widths: List of column widths in Excel units. If None, auto-sized
		styles: Dictionary defining styles for cells, rows, and columns
			- as returned by XLSXStyleBuilder.result
	Returns:
		BytesIO | None: BytesIO object containing the Excel file data if a new workbook was created, otherwise None

	"""
	column_widths = column_widths or []
	styles = styles or {}

	# creating workbook
	xlsx_file = None
	created_wb = False  # to know to close it later

	if wb is None:
		xlsx_file = BytesIO()
		options = {"constant_memory": True}

		if not styles:
			options["default_date_format"] = XLSXStyleBuilder.get_datetime_format()

		wb = xlsxwriter.Workbook(xlsx_file, options)
		created_wb = True

	ws = wb.add_worksheet(get_sanitized_sheet_name(sheet_name))

	# extract style components
	def _extract_ids(key: str) -> dict:
		return {k: tuple(v) for k, v in (styles.get(key) or {}).items() if v}

	style_registry: list[dict] = styles.get("styles") or []
	col_style_ids: dict[int, tuple[int, ...]] = _extract_ids("column_styles")
	row_style_ids: dict[int, tuple[int, ...]] = _extract_ids("row_styles")
	cell_style_ids: dict[tuple[int, int], tuple[int, ...]] = _extract_ids("cell_styles")

	styling_enabled = bool(col_style_ids or row_style_ids or cell_style_ids)

	if not styling_enabled:
		ws.set_row(0, cell_format=wb.add_format({"bold": True}))

	def resolve_style_ids(style_ids: tuple[int, ...]) -> dict:
		if len(style_ids) == 1:
			return style_registry[style_ids[0]]

		result = {}

		for sid in style_ids:
			result.update(style_registry[sid])
		return result

	@functools.cache
	def get_format(style_ids: tuple[int, ...]) -> Format:
		return wb.add_format(resolve_style_ids(style_ids))

	# set column widths
	for i, column_width in enumerate(column_widths):
		if column_width:
			ws.set_column(i, i, column_width)

	# column level styles
	for col_idx, style_ids in col_style_ids.items():
		ws.set_column(col_idx, col_idx, cell_format=get_format(style_ids))

	# row level styles (sorted because constant_memory mode requires writing rows in order)
	for row_idx, style_ids in sorted(row_style_ids.items()):
		ws.set_row(row_idx, cell_format=get_format(style_ids))

	# priority: column < row < cell (later in tuple = higher priority)
	cell_formats: dict[tuple[int, int], Format] = {}

	# process explicit cell styles
	for pos, cell_ids in cell_style_ids.items():
		row_idx, col_idx = pos
		col_ids = col_style_ids.get(col_idx, ())
		row_ids = row_style_ids.get(row_idx, ())

		cell_formats[pos] = get_format(col_ids + row_ids + cell_ids)

	# process row x column intersections (no explicit cell style)
	for row_idx, row_ids in row_style_ids.items():
		for col_idx, col_ids in col_style_ids.items():
			pos = (row_idx, col_idx)
			if pos not in cell_formats:
				cell_formats[pos] = get_format(col_ids + row_ids)

	# quick access for hot loop
	handle_html_content = sheet_name not in {"Data Import Template", "Data Export"}
	illegal_chars_search = ILLEGAL_CHARACTERS_RE.search
	illegal_chars_sub = ILLEGAL_CHARACTERS_RE.sub

	write = ws.write
	has_cell_formats = bool(cell_formats)
	get_cell_format = cell_formats.get

	for row_idx, row in enumerate(data):
		for col_idx, value in enumerate(row):
			if isinstance(value, str):
				if handle_html_content:
					value = handle_html(value)

				if illegal_chars_search(value):
					value = illegal_chars_sub("", value)

			cell_format = get_cell_format((row_idx, col_idx)) if has_cell_formats else None
			write(row_idx, col_idx, value, cell_format)

	if not created_wb:
		return

	wb.close()
	xlsx_file.seek(0)
	return xlsx_file


### Utilities ###
def get_sanitized_sheet_name(name: str) -> str:
	return INVALID_SHEET_NAME_RE.sub(" ", name)[:MAX_SHEET_NAME_LENGTH]


def handle_html(data: str) -> str:
	# return if no html tags found
	if "<" not in data or ">" not in data:
		return data

	h = unescape_html(data or "")

	try:
		value = html2text(h, strip_links=True, wrap=False)
	except Exception:
		# unable to parse html, send it raw
		return data

	return value.replace("  \n", ", ").replace("\n", " ").replace("# ", ", ")


def read_xlsx_file_from_attached_file(file_url=None, fcontent=None, filepath=None):
	if file_url:
		_file = frappe.get_doc("File", {"file_url": file_url})
		filename = _file.get_full_path()
	elif fcontent:
		filename = BytesIO(fcontent)
	elif filepath:
		filename = filepath
	else:
		return

	rows = []
	wb1 = load_workbook(filename=filename, data_only=True)
	ws1 = wb1.active
	for row in ws1.iter_rows():
		rows.append([cell.value for cell in row])
	return rows


def read_xls_file_from_attached_file(content):
	book = xlrd.open_workbook(file_contents=content)
	sheets = book.sheets()
	sheet = sheets[0]
	return [sheet.row_values(i) for i in range(sheet.nrows)]


def build_xlsx_response(data, filename, styles: dict | None = None):
	from frappe.desk.utils import provide_binary_file

	provide_binary_file(filename, "xlsx", make_xlsx(data, filename, styles=styles).getvalue())
